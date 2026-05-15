"""
RentCafe / Yardi multi-property apartment scraper.

Scrapes unit-level availability (unit number, beds, baths, sqft, rent,
availability date) from RentCafe-powered apartment sites and writes results
to apartments_<DATE>.xlsx (appending if the file already exists).

Properties covered (Northern Virginia cluster):
    - Windsor Kingstowne
    - Henley at Kingstowne
    - Contempo NOVA
    - Cameron Square
    - Springfield Crossing
    - Lerner Springfield Square
    - Park Place at Van Dorn
    - Vistas of Annandale
    - Woodside Apartments (Lorton)
    - Monticello of Falls Church
    - The Parliaments

Each RentCafe site exposes a `default.aspx` (or property domain root) page that
inline-renders a table per floorplan: `<table id="floorplanUnits...">` whose
`<tr class="fp-unit">` rows carry `data-unit-*` attributes plus an availability
cell with "Now" or a date. The wrapping `div.fp-item` carries `data-name`
(floorplan name), `data-beds`, `data-baths`, `data-size`, `data-rent`.
"""

import os
import re
import time
from datetime import datetime, date

import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth


# Each entry: (display_name, url)
# Prefer rentcafe.com mirrors — they all share the same RentCafe template.
PROPERTIES = [
    ("Windsor Kingstowne",
     "https://www.rentcafe.com/apartments/va/alexandria/windsor-kingstowne/default.aspx"),
    ("Henley at Kingstowne",
     "https://www.rentcafe.com/apartments/va/alexandria/henley-at-kingstown/default.aspx"),
    ("Contempo NOVA",
     "https://www.rentcafe.com/apartments/va/alexandria/contempo-nova/default.aspx"),
    ("Cameron Square",
     "https://www.rentcafe.com/apartments/va/alexandria/cameron-square/default.aspx"),
    ("Springfield Crossing",
     "https://www.rentcafe.com/housing/va/springfield/springfield-crossing/default.aspx"),
    ("Lerner Springfield Square",
     "https://www.rentcafe.com/apartments/va/springfield/springfield-square/default.aspx"),
    ("Park Place at Van Dorn",
     "https://www.rentcafe.com/apartments/va/alexandria/park-place-at-van-dorn1/default.aspx"),
    ("Vistas of Annandale",
     "https://www.rentcafe.com/apartments/va/annandale/vistas-of-annandale0/default.aspx"),
    ("Woodside Apartments",
     "https://www.rentcafe.com/apartments/va/lorton/woodside-0/default.aspx"),
    ("Monticello of Falls Church",
     "https://www.rentcafe.com/apartments/va/falls-church/monticello-falls-church-0/default.aspx"),
    ("The Parliaments",
     "https://www.rentcafe.com/apartments/va/annandale/parliaments/default.aspx"),
    ("Abbotts Run",
     "https://www.rentcafe.com/apartments/va/alexandria/abbotts-run-apartment-homes/default.aspx"),
]

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

HEADERS = ["Building", "Unit ID", "Floorplan", "Bedrooms", "Bathrooms",
           "Price", "Slashed Price", "Available Date", "Size (sq ft)",
           "Date Run", "Time Run"]


def _clean(s):
    return re.sub(r"\s+", " ", (s or "")).strip()


def parse_units(html, building_name, date_run, time_run):
    """Parse all unit rows from a RentCafe property page."""
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    seen_ids = set()

    # Each floorplan is a div.fp-item; nested table rows are tr.fp-unit
    for fp in soup.select("div.fp-item"):
        fp_name = _clean(fp.get("data-name", "")) or _clean(
            (fp.select_one(".fp-name") or {}).get_text(strip=True)
            if fp.select_one(".fp-name") else ""
        )
        fp_beds = _clean(fp.get("data-beds", ""))
        fp_baths = _clean(fp.get("data-baths", ""))

        for tr in fp.select("tr.fp-unit"):
            unit_id = _clean(tr.get("data-unit-name") or tr.get("data-unit-id"))
            if not unit_id or unit_id in seen_ids:
                continue
            seen_ids.add(unit_id)

            beds = _clean(tr.get("data-unit-beds") or fp_beds)
            baths = _clean(tr.get("data-unit-baths") or fp_baths)
            size_raw = _clean(tr.get("data-unit-size", ""))
            size = ""
            m = re.search(r"([\d,]+)", size_raw)
            if m:
                size = m.group(1).replace(",", "")

            rent = _clean(tr.get("data-unit-rent", ""))

            # Slashed/strikethrough rent (price drop)
            slashed = ""
            slash_el = tr.select_one(".strikethrough-pricing, s, del, .price-drop-original")
            if slash_el:
                slashed = _clean(slash_el.get_text())

            # Availability cell: typically the 3rd <td>; fall back to text scan.
            avail = ""
            tds = tr.find_all("td")
            if tds:
                # Skip the first td which usually holds rent
                for td in tds:
                    text = _clean(td.get_text(" "))
                    if not text:
                        continue
                    if re.search(r"\bNow\b|Available|\d{1,2}[/-]\d{1,2}|"
                                 r"Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec",
                                 text, re.I):
                        avail = text
                        break

            rows.append([
                building_name, unit_id, fp_name, beds, baths,
                rent, slashed, avail, size, date_run, time_run,
            ])

    return rows


def scrape_property(context, building_name, url):
    print(f"Opening: {building_name}")
    page = context.new_page()
    Stealth().apply_stealth_sync(page)
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        # Let any unit tables render
        try:
            page.wait_for_selector("div.fp-item, tr.fp-unit", timeout=15000)
        except Exception:
            print(f"  Warning: no fp-item found on {building_name}")
        # Small grace period for client-side rendering of unit rows
        page.wait_for_timeout(2500)
        html = page.content()
    finally:
        page.close()
    return html


FP_IDX = HEADERS.index("Floorplan")


def save_to_excel(all_rows, filename):
    """Append rows to today's apartments_<date>.xlsx (shared across all scrapers).
    Drops the Floorplan column when the existing file uses the legacy 10-column schema."""
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        has_floorplan = "Floorplan" in existing
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Units"
        ws.append(HEADERS)
        has_floorplan = True
    for row in all_rows:
        if not has_floorplan:
            row = [v for i, v in enumerate(row) if i != FP_IDX]
        ws.append(row)
    wb.save(filename)


def main():
    now = datetime.now()
    date_run = now.strftime("%m/%d/%Y")
    time_run = now.strftime("%H:%M:%S")

    all_rows = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        context = browser.new_context(
            user_agent=UA,
            locale="en-US",
            viewport={"width": 1280, "height": 900},
        )

        for i, (name, url) in enumerate(PROPERTIES):
            if i > 0:
                time.sleep(4)
            try:
                html = scrape_property(context, name, url)
                rows = parse_units(html, name, date_run, time_run)
                print(f"  {len(rows)} units")
                all_rows.extend(rows)
            except Exception as e:
                print(f"  ERROR scraping {name}: {e}")

        browser.close()

    output = f"apartments_{date.today()}.xlsx"
    save_to_excel(all_rows, output)
    print(f"\nDone. {len(all_rows)} RentCafe units written to {output}")


if __name__ == "__main__":
    main()

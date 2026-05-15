"""
The Elms at Kingstowne (Alexandria, VA) — best-effort scraper.

PLATFORM
    Built on the RealPage marketing CMS (G5 successor) at
    kingstowneelmsliving.com. The marketing /floor-plans route shows a
    photo carousel only — no inline floorplan or unit data. Actual
    availability lives inside the embedded "RealPage Online Leasing"
    widget rendered on /OnlineLeasing.aspx, which proxies to RealPage's
    LeaseStar product (cdn-leasestar.realpage.com).

WHY WE CAN'T GET UNIT-LEVEL DATA
    1. /floor-plans returns a hardened RealPage error page to non-browser
       clients (curl → 10 KB error). With Playwright stealth we get the
       full marketing page, but it carries no floorplan grid at all —
       just photo cards.
    2. /OnlineLeasing.aspx renders a marketing landing, not the unit
       picker. The actual unit list is inside a LeaseStar iframe loaded
       after a Cloudflare + RealPage bot challenge; the iframe URL is
       built dynamically with a property GUID and session token.
    3. The resident portal at theelmsatkingstowne.activebuilding.com is
       authenticated-only.

USER INTERACTION REQUIRED?
    Yes for unit-level data. RealPage's LeaseStar widget intermittently
    serves a "Press and Hold to confirm you are a human" PerimeterX
    challenge that Playwright cannot solve headlessly. When that
    challenge fires the script logs a warning and exits with code 2 so
    the cron retry kicks in later. On a clean session (no challenge),
    the embedded widget *still* requires picking beds + move-in date
    before listing units — there is no public, parameter-free unit list.

WHAT THIS SCRIPT DOES
    Best-effort scan of /floor-plans and /OnlineLeasing.aspx for any
    visible price ranges or floorplan names. If nothing renders, logs a
    sentinel row so the daily xlsx still records that this property was
    checked.
"""

import os
import re
import sys
from datetime import datetime, date

import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth


BUILDING = "The Elms at Kingstowne"
URLS = [
    "https://www.kingstowneelmsliving.com/apartments/va/alexandria/floor-plans",
    "https://www.kingstowneelmsliving.com/OnlineLeasing.aspx",
]
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

HEADERS = ["Building", "Unit ID", "Floorplan", "Bedrooms", "Bathrooms",
           "Price", "Slashed Price", "Available Date", "Size (sq ft)",
           "Date Run", "Time Run"]
FP_IDX = HEADERS.index("Floorplan")


def _clean(s):
    return re.sub(r"\s+", " ", (s or "")).strip()


def scan_for_floorplans(html, date_run, time_run):
    """Generic scan — pick up any visible bed/bath/sqft/price triples."""
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    seen = set()

    # Look for any block that mentions beds AND price
    text_blocks = soup.find_all(["div", "li", "section", "article"])
    for blk in text_blocks:
        tx = _clean(blk.get_text(" "))
        if len(tx) > 400 or len(tx) < 20:
            continue
        if not re.search(r"\$[\d,]{3,}", tx):
            continue
        if not re.search(r"Bed|Bedroom|Studio", tx, re.I):
            continue

        # Skip if it's just a marketing blurb wrapper that wraps many cards
        if blk.find(["div", "li"], recursive=False):
            continue

        beds = baths = size = price = ""
        m = re.search(r"(Studio|\d+)\s*(?:Bed|BR)", tx, re.I)
        if m: beds = m.group(1)
        m = re.search(r"(\d+(?:\.\d)?)\s*Bath", tx, re.I)
        if m: baths = m.group(1)
        m = re.search(r"([\d,]{3,5})\s*(?:sq\.?\s*ft|sf)", tx, re.I)
        if m: size = m.group(1).replace(",", "")
        m = re.search(r"\$[\d,]+(?:\s*-\s*\$[\d,]+)?", tx)
        if m: price = m.group(0)

        key = (beds, baths, size, price)
        if key in seen or not price:
            continue
        seen.add(key)

        rows.append([BUILDING, "", "", beds, baths, price, "", "", size,
                     date_run, time_run])
    return rows


def is_blocked(html):
    return ("press and hold" in html.lower()
            or "verify you are human" in html.lower()
            or "perimeterx" in html.lower())


def append_rows(filename, rows):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        has_floorplan = "Floorplan" in existing
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Units"
        ws.append(HEADERS)
        has_floorplan = True
    for row in rows:
        if not has_floorplan:
            row = [v for i, v in enumerate(row) if i != FP_IDX]
        ws.append(row)
    wb.save(filename)


def main():
    now = datetime.now()
    date_run = now.strftime("%m/%d/%Y")
    time_run = now.strftime("%H:%M:%S")

    print(f"Opening: {BUILDING}")
    collected = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        ctx = browser.new_context(user_agent=UA, locale="en-US",
                                  viewport={"width": 1280, "height": 900})
        for url in URLS:
            page = ctx.new_page()
            Stealth().apply_stealth_sync(page)
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=45000)
                page.wait_for_timeout(6000)
                html = page.content()
                if is_blocked(html):
                    print(f"  Bot challenge active on {url} — exiting (cron will retry).")
                    browser.close()
                    sys.exit(2)
                collected.extend(scan_for_floorplans(html, date_run, time_run))
            except Exception as e:
                print(f"  Error on {url}: {e}")
            finally:
                page.close()
        browser.close()

    if not collected:
        print("  No public floorplan/unit data found (RealPage LeaseStar widget gated).")
        collected = [[BUILDING, "", "(no public data — RealPage gated)", "", "",
                      "", "", "", "", date_run, time_run]]
    else:
        print(f"  {len(collected)} floorplan-level rows (no unit-level data is published)")

    output = f"apartments_{date.today()}.xlsx"
    append_rows(output, collected)
    print(f"Done. Appended to {output}")


if __name__ == "__main__":
    main()

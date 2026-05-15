"""
Cameron Square (Alexandria, VA) — best-effort scraper.

PLATFORM
    Bozzuto-managed property. Marketing site cameronsquare.com runs on a
    RentCafe-style "FloorplansV2" template (data-floorplan-name,
    data-floorplan-price, data-floorplan-sqft on each div.fp-container).
    The parent Bozzuto listing at bozzuto.com is a Next.js marketing page
    with no inline floorplan data.

WHY WE CAN'T GET UNIT-LEVEL DATA
    The public floorplans page exposes ONLY floorplan-level info:
    floorplan name, beds, baths, sqft range, and a price range like
    "$1,791 - $2,488". It does NOT publish individual apartment numbers,
    per-unit rents, or availability dates — those are gated behind a
    leasing-office contact form ("Contact Us" → leasing agent emails the
    unit list manually). There is no public AJAX endpoint that returns
    per-unit data the way RentCafe's fp-unit rows do.

USER INTERACTION REQUIRED FOR PER-UNIT DATA?
    Yes — humans only. The property does not publish per-unit pricing on
    the public web; tenants must request it from the leasing office.

WHAT THIS SCRIPT DOES
    Captures floorplan-level data so the daily xlsx still gets a row per
    floorplan (Unit ID is left blank, Price holds the range, Beds/Baths/
    Size are populated). This lets you track *floorplan* pricing changes
    over time even if you can't track individual units.
"""

import os
import re
import time
from datetime import datetime, date

import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth


BUILDING = "Cameron Square"
URL = "https://www.cameronsquare.com/floorplans"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

HEADERS = ["Building", "Unit ID", "Floorplan", "Bedrooms", "Bathrooms",
           "Price", "Slashed Price", "Available Date", "Size (sq ft)",
           "Date Run", "Time Run"]
FP_IDX = HEADERS.index("Floorplan")


def _clean(s):
    return re.sub(r"\s+", " ", (s or "")).strip()


def parse_floorplans(html, date_run, time_run):
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    seen = set()

    for card in soup.select("div.fp-container, div.fp-item"):
        # Pull data-* first (most reliable), fall back to visible text.
        fp_name = (card.get("data-floorplan-name")
                   or card.get("data-name")
                   or _clean(card.select_one(".fp-name").get_text() if card.select_one(".fp-name") else ""))
        if not fp_name:
            title = card.select_one(".card-title, h2, h3")
            if title:
                fp_name = _clean(title.get_text())
        fp_name = _clean(fp_name)
        if not fp_name or fp_name in seen:
            continue
        seen.add(fp_name)

        size_raw = card.get("data-floorplan-sqft") or card.get("data-floorplan-size") or ""
        size = ""
        m = re.search(r"[\d,]+", size_raw)
        if m:
            size = m.group(0).replace(",", "")
        else:
            # Visible "Sq. Ft." text in the card
            for el in card.select("p, span, div"):
                tx = el.get_text(" ", strip=True)
                m = re.search(r"([\d,]{3,5})\s*(?:sq\.?\s*ft|sf)", tx, re.I)
                if m:
                    size = m.group(1).replace(",", "")
                    break

        price = _clean(card.get("data-floorplan-price", ""))
        if not price:
            for el in card.select(".modal-fp-base-rent, .modal-fp-total-rent, .fp-price, .data-price, .price"):
                tx = _clean(el.get_text(" "))
                m = re.search(r"\$[\d,]+(?:\s*-\s*\$[\d,]+)?", tx)
                if m:
                    price = m.group(0)
                    break

        beds = ""
        baths = ""
        beds_baths_text = _clean(card.get_text(" "))
        m = re.search(r"(Studio|\d+)\s*(?:Bed|BR|Bedroom)", beds_baths_text, re.I)
        if m:
            beds = m.group(1)
        m = re.search(r"(\d+(?:\.\d)?)\s*Bath", beds_baths_text, re.I)
        if m:
            baths = m.group(1)

        rows.append([BUILDING, "", fp_name, beds, baths, price, "", "",
                     size, date_run, time_run])

    return rows


def append_rows(filename, rows):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        has_floorplan = "Floorplan" in existing
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Units"
        ws.append(HEADERS)
        has_floorplan = True
    for row in rows:
        if not has_floorplan:
            row = [v for i, v in enumerate(row) if i != FP_IDX]
        ws.append(row)
    wb.save(filename)


def main():
    now = datetime.now()
    date_run = now.strftime("%m/%d/%Y")
    time_run = now.strftime("%H:%M:%S")

    print(f"Opening: {BUILDING}")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        ctx = browser.new_context(user_agent=UA, locale="en-US",
                                  viewport={"width": 1280, "height": 900})
        page = ctx.new_page()
        Stealth().apply_stealth_sync(page)
        page.goto(URL, wait_until="domcontentloaded", timeout=45000)
        try:
            page.wait_for_selector("div.fp-container, div.fp-item", timeout=15000)
        except Exception:
            print("  Warning: floorplan cards not detected.")
        page.wait_for_timeout(2500)
        html = page.content()
        browser.close()

    rows = parse_floorplans(html, date_run, time_run)
    print(f"  {len(rows)} floorplans (floorplan-level only — see header comment)")
    output = f"apartments_{date.today()}.xlsx"
    append_rows(output, rows)
    print(f"Done. Appended to {output}")


if __name__ == "__main__":
    main()

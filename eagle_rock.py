"""
Eagle Rock Apartments at West Springfield — best-effort scraper.

PLATFORM
    G5 Marketing Cloud frontend (eaglerockproperties.com) backed by
    RentManager (eaglerock.twa.rentmanager.com). The "Floorplan &
    Availability" page is at /apartments/va/springfield/burling-wood-dr/
    floorplan-availability.

WHY WE CAN'T GET UNIT-LEVEL DATA
    The G5 page renders no inline unit rows. Live availability is gated
    behind a BetterBot conversational widget that posts to
    https://westspring2.engine.betterbot.com/api/details and only reveals
    units after the user picks bed-count / move-in-date / budget. The
    RentManager tenant portal at eaglerock.twa.rentmanager.com is
    authenticated and shows nothing pre-login. Apartments-com style
    syndication is the only place per-unit pricing currently appears.

USER INTERACTION REQUIRED?
    Effectively yes — getting per-unit pricing requires walking through
    the BetterBot chatbot turn by turn (select beds → date → email).
    The chatbot will also email/SMS the lead to a leasing agent, which is
    not appropriate for an automated daily scrape.

WHAT THIS SCRIPT DOES
    Best-effort: open the floorplan-availability page and capture any
    floorplan-level summaries it manages to render (beds, baths, sqft).
    If nothing renders (the page currently produces a marketing layout
    with no public floorplan grid), the script logs that condition and
    appends a single "no public data" sentinel row so the daily xlsx
    shows the property was checked but skipped.
"""

import os
import re
from datetime import datetime, date

import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth


BUILDING = "Eagle Rock Apartments at West Springfield"
URL = ("https://www.eaglerockproperties.com/apartments/va/springfield/"
       "burling-wood-dr/floorplan-availability")
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

HEADERS = ["Building", "Unit ID", "Floorplan", "Bedrooms", "Bathrooms",
           "Price", "Slashed Price", "Available Date", "Size (sq ft)",
           "Date Run", "Time Run"]
FP_IDX = HEADERS.index("Floorplan")


def _clean(s):
    return re.sub(r"\s+", " ", (s or "")).strip()


def parse_floorplans(html, date_run, time_run):
    """Scan for any visible floorplan or unit info on the G5 page."""
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    # G5 floorplan cards typically have classes like 'fp', 'floorplan', or
    # 'available-floorplan'. We also check for itemprop="floorplan" microdata.
    candidates = soup.select(
        '[itemtype*="FloorPlan"], .floorplan, .fp, [class*="floorplan-card"], '
        '[class*="availability-card"], [class*="fp-card"]'
    )

    for card in candidates:
        text = _clean(card.get_text(" "))
        fp_name = ""
        h = card.select_one("h1, h2, h3, h4, [class*='title'], [class*='name']")
        if h:
            fp_name = _clean(h.get_text())

        beds = baths = size = price = avail = ""
        m = re.search(r"(Studio|\d+)\s*(?:Bed|BR)", text, re.I)
        if m: beds = m.group(1)
        m = re.search(r"(\d+(?:\.\d)?)\s*Bath", text, re.I)
        if m: baths = m.group(1)
        m = re.search(r"([\d,]{3,5})\s*(?:sq\.?\s*ft|sf)", text, re.I)
        if m: size = m.group(1).replace(",", "")
        m = re.search(r"\$[\d,]+(?:\s*-\s*\$[\d,]+)?", text)
        if m: price = m.group(0)
        m = re.search(r"Available\s+([A-Za-z]+\s+\d+|Now|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", text, re.I)
        if m: avail = m.group(1)

        if fp_name or price or beds:
            rows.append([BUILDING, "", fp_name, beds, baths, price, "",
                         avail, size, date_run, time_run])

    return rows


def append_rows(filename, rows):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        has_floorplan = "Floorplan" in existing
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Units"
        ws.append(HEADERS)
        has_floorplan = True
    for row in rows:
        if not has_floorplan:
            row = [v for i, v in enumerate(row) if i != FP_IDX]
        ws.append(row)
    wb.save(filename)


def main():
    now = datetime.now()
    date_run = now.strftime("%m/%d/%Y")
    time_run = now.strftime("%H:%M:%S")

    print(f"Opening: {BUILDING}")
    html = ""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        ctx = browser.new_context(user_agent=UA, locale="en-US",
                                  viewport={"width": 1280, "height": 900})
        page = ctx.new_page()
        Stealth().apply_stealth_sync(page)
        try:
            page.goto(URL, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(8000)
            html = page.content()
        except Exception as e:
            print(f"  Network error: {e}")
        browser.close()

    rows = parse_floorplans(html, date_run, time_run) if html else []
    if not rows:
        print("  No public floorplan/unit data found (gated by BetterBot widget).")
        print("  Logging sentinel row so daily xlsx records the attempt.")
        rows = [[BUILDING, "", "(no public data — BetterBot gated)", "", "",
                 "", "", "", "", date_run, time_run]]
    else:
        print(f"  {len(rows)} floorplan-level rows (no unit-level data is published)")

    output = f"apartments_{date.today()}.xlsx"
    append_rows(output, rows)
    print(f"Done. Appended to {output}")


if __name__ == "__main__":
    main()

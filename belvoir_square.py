"""
Belvoir Square (Fort Belvoir, VA) — best-effort scraper.

PLATFORM
    Five Star Living / McKinney Properties branded site backed by a Yardi
    RentCafe "Five Star" skin. Live URL pattern:
    /fort-belvoir/belvoir-square/conventional/

WHY UNIT-LEVEL DATA IS LIMITED
    The conventional/ landing page renders floorplan-level rows
    (div.fp-col grid: image / beds-bath / sq-feet / rent / special /
    action). Individual unit numbers / per-unit rents are NOT inlined —
    each floorplan links to a per-floorplan detail page at
    /floorplans/fort-belvoir-VA/belvoir-square/<plan-slug>-<id>-1/ where
    sometimes individual units appear, but as of last check those
    sub-pages only echo the same min-price/sqft range, not unit numbers.

    The site sits behind a Cloudflare bot wall — default curl returns 403.
    Playwright stealth (same approach used by the working scrapers in
    this repo) gets through.

USER INTERACTION REQUIRED?
    No interactive CAPTCHA on a typical visit (Cloudflare uses a
    transparent JS challenge that Playwright passes). If the site ever
    escalates to a managed challenge, the page will show "Verify you are
    human" and Playwright will hang on it — the script prints a warning
    and exits with code 2 in that case so the cron retry logic kicks in.

WHAT THIS SCRIPT DOES
    Parses div.fp-col-wrapper rows on /conventional/ to extract per-
    floorplan: name (fp-col-title), beds/baths (fp-col bed-bath), sqft
    (fp-col sq-feet), and current rent range (fp-col rent).
"""

import os
import re
import sys
from datetime import datetime, date

import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth


BUILDING = "Belvoir Square"
URL = "https://www.belvoirsquare.com/fort-belvoir/belvoir-square/conventional/"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")

HEADERS = ["Building", "Unit ID", "Floorplan", "Bedrooms", "Bathrooms",
           "Price", "Slashed Price", "Available Date", "Size (sq ft)",
           "Date Run", "Time Run"]
FP_IDX = HEADERS.index("Floorplan")


def _clean(s):
    return re.sub(r"\s+", " ", (s or "")).strip()


def parse_floorplans(html, date_run, time_run):
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    seen = set()

    # Each floorplan is a li.fp-group-item containing fp-col cells.
    for item in soup.select("li.fp-group-item"):
        cells = {}
        for col in item.select("div.fp-col"):
            cls = " ".join(col.get("class", []))
            txt = _clean(col.get_text(" "))
            for key in ("bed-bath", "sq-feet", "rent", "special", "action"):
                if key in cls:
                    cells.setdefault(key, txt)

        # Floorplan name: prefer the floorplan-detail URL slug since it is
        # always unique (img alt is reused across most plans).
        fp_name = ""
        a = item.find("a", href=re.compile(r"/floorplans/"))
        if a:
            m = re.search(r"/([^/]+)-\d+-\d+/?$", a.get("href", ""))
            if m:
                fp_name = m.group(1).replace("-", " ").title()
        if not fp_name:
            img = item.find("img")
            if img:
                fp_name = _clean(img.get("alt", ""))
        if not fp_name or fp_name in seen:
            continue
        seen.add(fp_name)

        bb = cells.get("bed-bath", "")
        beds = baths = ""
        m = re.search(r"(Studio|\d+)\s*(?:Bed|BR|/)", bb, re.I)
        if m: beds = m.group(1)
        m = re.search(r"(\d+(?:\.\d)?)\s*(?:Bath|ba)", bb, re.I)
        if m: baths = m.group(1)

        size = ""
        m = re.search(r"[\d,]+", cells.get("sq-feet", ""))
        if m: size = m.group(0).replace(",", "")

        price = ""
        m = re.search(r"\$[\d,]+(?:\s*-\s*\$[\d,]+)?", cells.get("rent", ""))
        if m: price = m.group(0)

        # Availability date from the action cell, e.g. "Available Jun 24, 2026"
        avail = ""
        m = re.search(r"Available\s+([A-Za-z]+\s+\d{1,2},?\s+\d{4}|Now)",
                      cells.get("action", ""), re.I)
        if m: avail = m.group(1)

        rows.append([BUILDING, "", fp_name, beds, baths, price,
                     "", avail, size, date_run, time_run])

    return rows


def append_rows(filename, rows):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        has_floorplan = "Floorplan" in existing
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Units"
        ws.append(HEADERS)
        has_floorplan = True
    for row in rows:
        if not has_floorplan:
            row = [v for i, v in enumerate(row) if i != FP_IDX]
        ws.append(row)
    wb.save(filename)


def main():
    now = datetime.now()
    date_run = now.strftime("%m/%d/%Y")
    time_run = now.strftime("%H:%M:%S")

    print(f"Opening: {BUILDING}")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, channel="chrome")
        ctx = browser.new_context(user_agent=UA, locale="en-US",
                                  viewport={"width": 1280, "height": 900})
        page = ctx.new_page()
        Stealth().apply_stealth_sync(page)
        page.goto(URL, wait_until="domcontentloaded", timeout=45000)
        try:
            page.wait_for_selector("li.fp-group-item, div.fp-col", timeout=15000)
        except Exception:
            html = page.content()
            if "Verify you are human" in html or "challenge-platform" in html:
                print("  Cloudflare managed challenge active — manual interaction "
                      "required. Exiting so cron retry can try again later.")
                browser.close()
                sys.exit(2)
            print("  Warning: floorplan rows not detected.")
        page.wait_for_timeout(2500)
        html = page.content()
        browser.close()

    rows = parse_floorplans(html, date_run, time_run)
    print(f"  {len(rows)} floorplans (floorplan-level only — see header comment)")
    output = f"apartments_{date.today()}.xlsx"
    append_rows(output, rows)
    print(f"Done. Appended to {output}")


if __name__ == "__main__":
    main()
